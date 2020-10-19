# -*- coding: utf-8 -*-
from datetime import datetime, timedelta
from os import listdir
from openpyxl import Workbook
from json import loads

from sys import path
path.append('/simepar/hidro/COPEL/SISPSHI2/Bibliotecas/')
from admin import dataexec

print('\n ~/SISPSHI2/Site_Update/7-planilhas.py')

dirSite = '/simepar/hidro/webdocs/sispshi2/dados/'
dirPlan = '/simepar/hidro/webdocs/sispshi2/planilhas/'
jsons = sorted(listdir(dirSite))
ref = [None, None, None, None, None, None, None, None, None] #cmbobs, qobs, cmbprv, qprev<1,2,3,4>
dtref = dataexec(['datareferencia'])

for bac in range(1,22):
    
    cab = ['Chuva Media Observada','Vazao Observada',]
    dados = {}
    
    #Inserindo dados de chuva media na bacia observada
    arq = open(dirSite + str('cmb%2.2i.json' % bac), 'r')
    xy = loads(arq.read())
    for i in range(len(xy)):
        dt = datetime.strptime(xy[i][0], '%Y-%m-%d %H:%M:%S')
        dados[dt] = ref[:]
        dados[dt][0] = xy[i][1]
    arq.close()
            
    #Inserindo dados da vazao observada
    arq = open(dirSite + str('vaz%2.2i.json' % bac), 'r')
    xy = loads(arq.read())
    for i in range(len(xy)):
        dt = datetime.strptime(xy[i][0], '%Y-%m-%d %H:%M:%S')
        try:
            dados[dt][1] = xy[i][1]
        except KeyError:
            dados[dt] = ref[:]
            dados[dt][1] = xy[i][1]
    arq.close()
    
    #Previsão de chuva média
    cab.append('Chuva Prevista')
    arq = open(dirSite + str('prevcmb%2.2i.json' % bac), 'r')
    xy = loads(arq.read())
    for i in range(len(xy)):
        dt = datetime.strptime(xy[i][0], '%Y-%m-%d %H:%M:%S')
        if dt < dtref: continue
        try:
            dados[dt][2] = xy[i][1]
        except KeyError:
            dados[dt] = ref[:]
            dados[dt][2] = xy[i][1]
    arq.close()

    #Previsões de vazão
    aux = str('prevaz%2.2i' % bac)
    lista = [dirSite+js for js in jsons if js[0:8] == aux]
    iref = 3
    for item in lista:
        cab.append('Vazao Prev. '+item[-9:-5])
        arq = open(item, 'r')
        xy = loads(arq.read())
        for i in range(len(xy)):
            dt = datetime.strptime(xy[i][0], '%Y-%m-%d %H:%M:%S')
            if dt < dtref: continue
            try:
                dados[dt][iref] = xy[i][1]
            except KeyError:
                dados[dt] = ref[:]
                dados[dt][iref] = xy[i][1]
            #except IndexError:
                #print arq.name
                #print dt, len(dados[dt]), iref
                #print i, len(xy)
                #exit()
        arq.close()
        iref += 1

    #Inicializando arquivo.xlsx
    nCols = len(cab)
    wb = Workbook()
    ws = wb.active
    for nomeCol in ['A','B','C','D','E','F','G','H','I','J']:
        ws.column_dimensions[nomeCol].width = 19
    
    #Gravando cabeçalho
    ws.cell(row = 1, column = 1).value = 'Horario'
    for i in range(nCols):
        ws.cell(row = 1, column = i+2).value = cab[i]
    
    #Gravando dados na planilha
    datas = sorted(dados.keys())
    for i in range(len(datas)):
        dt = datas[i]
        ws.cell(row = i+2, column = 1).value = dt
        for j in range(nCols):
            if dados[dt][j] != None:
                ws.cell(row = i+2, column = j+2).value = dados[dt][j]
    
    #Salvando planilha
    wb.save(dirPlan + str('dados_hora_b%2.2i.xlsx' % bac))

print(' > Planilhas atualizadas!\n')